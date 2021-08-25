# Concatenates data from excel workbooks in the 'Collection date updated in Netsuite' folder to the master.xlsx in the 'Requests in Transit' folder.
# Then it moves the excel files to the 'Request in Transit' folder.

# The program uses the following modules
import os 
import shutil 
import openpyxl
from openpyxl.utils.cell import column_index_from_string, get_column_letter, coordinate_from_string


# os.chdir('/Volumes/GoogleDrive/Shared drives/Foreign Purchasing/Operations/Receiving Lists/Requests pending confirmation/Collection date updated in NetSuite')
path = '/Volumes/GoogleDrive/Shared drives/Foreign Purchasing/Operations/Receiving Lists/Requests pending confirmation/Collection date updated in NetSuite'

# Loop through the workbooks in the directory
for filename in os.listdir(path):
    sourceWorkbook = (os.path.join(path, filename))
    print('adding ' + filename + ' to master sheet')

    wb1 = openpyxl.load_workbook(sourceWorkbook)
    ws1 = wb1.active
    mr = ws1.max_row
    mc = ws1.max_column
    print(filename + ' has ' + str(mr) + ' rows')
    rowsToBeAdded = mr - 2
    print(str(rowsToBeAdded) + ' rows will be added to master.xlsx')

    wb2 = openpyxl.load_workbook('/Volumes/GoogleDrive/Shared drives/Foreign Purchasing/Operations/Receiving Lists'
                                '/Requests in Transit/master.xlsx')
    ws2 = wb2.active
    mr2 = ws2.max_row
    mc2 = ws2.max_column

    # Counter for row number of master excel sheet 
    number = mr2
    # Loop through the number of rows in source sheet 
    count = 0
    for i in range(2, mr):
        count += 1
        number += 1
        #Loop through the number of columns in source sheet 
        for j in range(1, mc +1):
            # Value from source worksheet 
            c = ws1.cell(row=i, column=j)
            # Assign value from source worksheet to destination sheet 
            ws2.cell(row= number, column = j).value = c.value
            # Save the new value to master(destination) sheet 
            wb2.save('/Volumes/GoogleDrive/Shared drives/Foreign Purchasing/Operations/Receiving Lists'
                                '/Requests in Transit/master.xlsx')
        print('.......')
    print(str(count) + ' rows have been added')    

# Move workbooks from one directory to another
source = '/Volumes/GoogleDrive/Shared drives/Foreign Purchasing/Operations/Receiving Lists/Requests pending confirmation/Collection date updated in NetSuite/'
dest = '/Volumes/GoogleDrive/Shared drives/Foreign Purchasing/Operations/Receiving Lists/Requests in Transit/'
for filename in os.listdir('/Volumes/GoogleDrive/Shared drives/Foreign Purchasing/Operations/Receiving Lists/Requests pending confirmation/Collection date updated in NetSuite'):
    shutil.move(source+filename, dest)
print('All files moved to Requests in Transit')  
